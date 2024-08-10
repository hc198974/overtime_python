# -*- coding: utf-8 -*-
import calendar
import datetime
import itertools
import tkinter
import tkinter.simpledialog
import requests
from lxml import etree
from openpyxl import load_workbook
import win32com.client
import time


class Crili(object):
    """
    万年日历接口数据抓取
    Params:year 四位数年份字符串
    """

    def __init__(self, year, month):
        self.year = year
        self.month = month

    def parseHTML(self):
        """页面解析"""
        global weekday
        url = 'https://wannianrili.bmcx.com/ajax/'
        s = requests.session()
        headers = {
            'Host': 'wannianrili.bmcx.com',
            'Connection': 'keep-alive',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36',
            'Accept': '*/*',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Referer': 'https://wannianrili.51240.com/',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        }
        result = {}

        c = calendar.monthrange(self.year, self.month)[1]
        s = requests.session()
        payload = {'q': str(self.year) + '-' + str(self.month)}
        response = s.get(url, headers=headers, params=payload)
        element = etree.HTML(response.text)
        html = element.xpath('//div[@class="wnrl_riqi"]')

        # 获取节点属性
        for i in range(c):
            item = html[i].xpath('./a')[0].attrib
            if item['id'] == 'wnrl_riqi_id_' + str(i):
                if 'class' in item:
                    temp = datetime.datetime(self.year, self.month, i + 1)
                    if item['class'] == 'wnrl_riqi_xiu':
                        weekday = 3
                    elif item['class'] == 'wnrl_riqi_mo':
                        weekday = 2
                    elif item['class'] == 'wnrl_riqi_ban':
                        weekday = 1.5
                else:
                    temp = datetime.datetime(self.year, self.month, i + 1)
                    if temp.weekday() > 4:
                        weekday = 2
                    else:
                        weekday = 1.5

                result[temp.strftime('%Y%m%d')] = weekday
        return (result)


class Cwindow(object):
    def __init__(self):
        self.month = datetime.datetime.now().month - 1

    def set_win_center(self, root, curWidth='', curHight=''):
        '''
    设置窗口大小，并居中显示
    param root:主窗体实例
    param curWidth:窗口宽度，非必填，默认200
    return:无
    '''
        if not curWidth:
            '''获取窗口宽度，默认200'''
            curWidth = root.winfo_width()
        if not curHight:
            '''获取窗口高度，默认200'''
            curHight = root.winfo_height()

        # 获取屏幕宽度和高度
        scn_w, scn_h = root.maxsize()

        # 计算中心坐标
        cen_x = (scn_w - curWidth) / 2
        cen_y = (scn_h - curHight) / 2

        # 设置窗口初始大小和位置
        size_xy = '%dx%d+%d+%d' % (curWidth, curHight, cen_x, cen_y)
        root.geometry(size_xy)

    def askName(self):
        # 获取字符串（标题，提示，初始值）
        name = tkinter.simpledialog.askstring(
            title='获取信息', prompt='请输入姓名：', initialvalue='韩超')
        self.name = name

    def askMonth(self):
        month = tkinter.simpledialog.askinteger(
            title='获取月份', prompt='请输入月份', initialvalue=datetime.datetime.now().month - 1)
        self.month = month

    def clearSheet(self):
        # 对汇总表数据进行清理
        Cmacro().dealData()

    def shutDown(self):
        root.destroy()

    def createWindow(self):
        global root
        # 创建主窗口
        root = tkinter.Tk()
        # 设置窗口大小
        root.resizable(False, False)
        root.title('加班')
        root.update()
        self.set_win_center(root, 300, 150)
        # 添加按钮
        # btn1 = tkinter.Button(root, text='获取用户名', command=self.askName)
        # btn1.pack(expand='yes')
        btn2 = tkinter.Button(root, text='获取月份', command=self.askMonth)
        btn2.pack(expand='yes')
        btn4 = tkinter.Button(root, text='清理数据', command=self.clearSheet)
        btn4.pack(expand='yes')
        btn3 = tkinter.Button(root, text='开始计算', command=self.shutDown)
        btn3.pack(expand='yes')
        # 加入消息循环
        root.mainloop()


class Cmacro():
    def __init__(self) -> None:
        self.path = 'C:\\Users\\Administrator\\Desktop\\overtime_python_slow\\原始数据.xlsm'

    def dealData(self):
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        wb = excel.Workbooks.Open(self.path)
        print('START')
        excel.Application.Run("deleteRow")
        wb.SaveAs(r'C:\Users\Administrator\Desktop\overtime_python_slow\计算结果.xlsx', FileFormat=51,
                  ConflictResolution=2)
        wb.Close()
        print('END')


class Count(object):
    def __init__(self, name, month, result):
        self.fpath = '计算结果.xlsx'
        # 节假日接口(工作日对应结果为 0, 休息日对应结果为 1, 节假日对应的结果为 2 )
        # server_url = "http://www.easybots.cn/api/holiday.php?d="
        self.server_url = "http://tool.bitefu.net/jiari/?d="
        self.wb = load_workbook(filename=self.fpath)
        self.ws = self.wb['汇总表']
        self.ws2 = self.wb['中干']
        self.name = name.value
        self.month = month
        self.dict = {}
        self.weekday = {}
        self.workday = {}
        self.holiday = {}
        self.cash = {}
        self.hour = 0
        self.result = result

    def getUrl(self):
        try:
            for m in self.result:
                if result[m] == 1.5:
                    self.workday[m] = 0
                elif result[m] == 2:
                    self.weekday[m] = 1
                elif result[m] == 3:
                    self.holiday[m] = 2

        except ConnectionResetError as e:
            print('远程主机发生错误' + e)

    # 调整表里的加班小时数
    def changeHour(self):
        temp17 = datetime.datetime.strptime('17:30', "%H:%M")
        temp18 = datetime.datetime.strptime('18:00', '%H:%M')
        temp12 = datetime.datetime.strptime('12:00', "%H:%M")
        temp13 = datetime.datetime.strptime('13:00', "%H:%M")
        temp8 = datetime.datetime.strptime('8:00', "%H:%M")
        self.getUrl()
        for x in self.ws.rows:
            if x[0].value == self.name:
                if x[4].value == self.month:
                    temp = x[1].value.strftime("%Y%m%d")
                    time1 = x[2].value
                    time2 = x[3].value
                    if time1 != '' and time2 != '' and time1 != None and time2 != None:
                        if time2 > time1:
                            time1 = datetime.datetime.strptime(time1, "%H:%M")
                            time2 = datetime.datetime.strptime(time2, "%H:%M")

                            # 工作日
                            if temp in self.workday:
                                if time2 > temp18:
                                    self.hour = (time2 - temp17).seconds

                            if self.hour > 0:
                                x[7].value = round(self.hour / 3600, 2)
                                s = x[1].value.strftime("%Y%m%d")
                                self.dict[s] = x[7].value
                                x[5].value = "工作日"
                                self.hour = 0
                            else:
                                x[7].value = 0
                                x[5].value = "工作日"
                                self.hour = 0

                            # 周末和节假日
                            if temp in self.weekday or temp in self.holiday:
                                if time1 > temp8:
                                    if temp12 < time1 < temp13:
                                        time1 = temp12
                                else:
                                    time1 = temp8

                                if temp12 < time2 < temp13:
                                    time2 = temp13
                                else:
                                    pass

                                if time2 <= temp12:
                                    self.hour = time2 - time1 - \
                                                datetime.timedelta(hours=0.5)
                                if time2 >= temp13:
                                    if time1 <= temp12:
                                        self.hour = time2 - time1 - \
                                                    datetime.timedelta(hours=1.5)
                                    else:
                                        self.hour = time2 - time1 - \
                                                    datetime.timedelta(hours=0.5)

                                if self.hour.days == 0:
                                    x[7].value = round(
                                        self.hour.seconds / 3600, 2)
                                    s = x[1].value.strftime("%Y%m%d")
                                    self.dict[s] = x[7].value
                                    x[5].value = "节假日"
                                    self.hour = 0
                                else:
                                    x[7].value = 0
                                    x[5].value = "节假日"
                                    self.hour = 0

        self.wb.save('计算结果.xlsx')

    def setConvert(self):
        # 写入是转加班费self.cash还是转串休self.dicts
        for x in self.ws.rows:
            if x[0].value == self.name:
                for y in self.cash.keys():
                    if x[1].value.strftime("%Y%m%d") == y:
                        x[6].value = "转加班费"

                for y in self.dict.keys():
                    if x[1].value.strftime("%Y%m%d") == y:
                        x[6].value = "转串休"
        self.wb.save('计算结果.xlsx')

    def setContents(self, sum, sum_chuan_xiu):
        rng = self.ws2['C2':'AG2']
        for x in rng:
            for y in x:
                for z in self.dict:
                    if y.value.strftime("%Y%m%d") == z:
                        self.ws2.cell(
                            row=name.row, column=y.column).value = self.dict[z]
        self.ws2.cell(row=name.row, column=34).value = sum
        self.ws2.cell(row=name.row, column=35).value = sum_chuan_xiu
        self.wb.save('计算结果.xlsx')

    def jiSuan(self):
        # 获得URL
        self.changeHour()
        dict_1, dict_2, dict_3 = {}, {}, {}
        for x in self.dict:
            if self.result[x] == 1.5:
                dict_1.update({x: self.dict[x]})
            elif self.result[x] == 2:
                dict_2.update(({x: self.dict[x]}))
            elif self.result[x] == 3:
                dict_3.update(({x: self.dict[x]}))

        remainder = 36
        if sum(list(self.dict.values())) > 36:
            for p in [dict_3, dict_2, dict_1]:
                if len(p) > 0:
                    combine = []
                    for r in range(1, len(p) + 1):
                        combinations = list(itertools.combinations(p, r))
                        for x in combinations:
                            combine.append(x)

                    temp = {}
                    smax = 0
                    total = 0
                    for m in combine:
                        for n in m:
                            total += self.dict[n]

                        if total <= remainder:
                            if smax < total:
                                smax = total
                                temp.clear()
                                for y in m:
                                    temp.update({y: self.dict[y]})
                                total = 0
                            else:
                                total = 0
                        else:
                            total = 0

                    self.cash.update(temp)
                    remainder = remainder - sum(list(self.cash.values()))
        else:
            self.cash = self.dict.copy()

        print('总数据一览：', self.dict)
        print('加班数合计：', round(sum(list(self.dict.values())), 2))
        print('转加班小时：', round(sum(list(self.cash.values())), 2))
        sum_chuan_xiu = round(sum(list(self.dict.values())) -
                              sum(list(self.cash.values())), 2)
        print('转串休小时：', sum_chuan_xiu)
        # 先清空单元格
        for row in self.ws2.iter_rows(min_row=name.row, max_row=name.row, min_col=3, max_col=35):
            for cell in row:
                cell.value = None
        self.setContents(sum(list(self.dict.values())), sum_chuan_xiu)
        print("转加班费：", sorted(self.cash.keys()))
        for k in self.cash.keys():
            self.dict.pop(k)
        print('转串休假：', sorted(self.dict.keys()))
        self.setConvert()


start = time.perf_counter()
cw = Cwindow()
cw.createWindow()
# 获得工作日和节假日
result = Crili(2024, cw.month).parseHTML()

wb = load_workbook(filename='原始数据.xlsm')
ws = wb['中干']
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=2):
    for name in row:
        if name.value is not None:
            print(name.value, cw.month, '月')
            ji = Count(name, cw.month, result)
            ji.jiSuan()
        else:
            break

end = time.perf_counter()
print("运行时间：", end - start)
